#!/usr/bin/env python3
"""
Regenerate NSSU_Common_Practice_Library_*.json from the matching .xlsx.

This script mirrors `parseLibraryWorkbook` in index.html exactly. Run it
whenever the .xlsx changes so the .json stays in sync.

Usage:
    python scripts/regenerate_library_json.py [path_to_xlsx]

Defaults to NSSU_Common_Practice_Library_v3_2.xlsx in the repo root.

Requires: openpyxl  (pip install openpyxl)
"""
import datetime
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

DEFAULT_XLSX = "NSSU_Common_Practice_Library_v3_2.xlsx"
VERSION = "3.2"  # bump when the spreadsheet version changes

SKIP_SHEETS = {"Instructions", "Engine_Logic", "Base_Behavior"}
RULE_SHEET_TO_KEY = {
    "Pressure_Overrides":     "pressureOverrides",
    "Process_Overrides":      "processOverrides",
    "PPE_Rules":              "ppeRules",
    "Data_Rules":             "dataRules",
    "Focal_Point_Overrides":  "focalOverrides",
}


def cell_str(value) -> str:
    return str(value).strip() if value is not None else ""


def main() -> int:
    xlsx_path = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)

    result = {
        "version": VERSION,
        "reworkAssociations": {},
        "pressureOverrides": {},
        "processOverrides": {},
        "ppeRules": {},
        "dataRules": {},
        "focalOverrides": {},
        "parsedAt": datetime.datetime.now(datetime.UTC).isoformat(),
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        sheet = wb[sheet_name]

        # Rule sheets: col A = condition/trigger, col B = rule
        if sheet_name in RULE_SHEET_TO_KEY:
            target = result[RULE_SHEET_TO_KEY[sheet_name]]
            for row in range(4, sheet.max_row + 1):
                condition = cell_str(sheet.cell(row, 1).value)
                rule      = cell_str(sheet.cell(row, 2).value)
                if condition and rule:
                    target[condition] = rule
            continue

        # Defect tab: tab name (lowercased, " - " → " / ") is the defect key
        defect_key = sheet_name.replace(" - ", " / ").lower()
        assoc = {}
        for row in range(4, sheet.max_row + 1):
            tool = cell_str(sheet.cell(row, 1).value).lower()
            lang = cell_str(sheet.cell(row, 2).value)
            if tool and lang and tool != "tool name":
                assoc[tool] = lang
        # Always create the defect entry, even if empty — matches the app
        result["reworkAssociations"][defect_key] = assoc

    output_path = xlsx_path.with_suffix(".json")
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    defect_count = len(result["reworkAssociations"])
    assoc_count  = sum(len(t) for t in result["reworkAssociations"].values())
    empty        = sum(1 for t in result["reworkAssociations"].values() if not t)
    rule_count   = sum(len(result[k]) for k in RULE_SHEET_TO_KEY.values())

    print(f"Wrote {output_path}")
    print(f"  version:           {result['version']}")
    print(f"  defects:           {defect_count}  ({empty} intentionally empty)")
    print(f"  associations:      {assoc_count}")
    print(f"  rule entries:      {rule_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
